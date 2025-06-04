"""
Loan Scenario Analysis Tool
Creates a beautiful Excel workbook for analyzing mortgage refinancing scenarios.
"""

import pandas as pd
import numpy as np
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule


class LoanScenarioAnalyzer:
    def __init__(self, apartment_loan=555000, apartment_rate=5.49, apartment_term=30,
                 investment_loan=220000, investment_rate=6.49, investment_term=30):
        self.apartment_loan = apartment_loan
        self.apartment_rate = apartment_rate
        self.apartment_term = apartment_term
        self.investment_loan = investment_loan
        self.investment_rate = investment_rate
        self.investment_term = investment_term
    
    def calculate_monthly_payment(self, principal, annual_rate, years):
        """Calculate monthly payment using PMT formula"""
        if annual_rate == 0:
            return principal / (years * 12)
        
        monthly_rate = annual_rate / 100 / 12
        num_payments = years * 12
        payment = principal * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)
        return payment
    
    def calculate_total_interest(self, principal, monthly_payment, years):
        """Calculate total interest over loan term"""
        return (monthly_payment * years * 12) - principal
    
    def create_amortization_schedule(self, principal, annual_rate, years, start_month=1):
        """Create amortization schedule"""
        monthly_rate = annual_rate / 100 / 12
        monthly_payment = self.calculate_monthly_payment(principal, annual_rate, years)
        
        schedule = []
        balance = principal
        
        for month in range(int(years * 12)):
            interest_payment = balance * monthly_rate
            principal_payment = monthly_payment - interest_payment
            balance -= principal_payment
            
            schedule.append({
                'Payment': month + 1,
                'Beginning_Balance': balance + principal_payment,
                'Payment_Amount': monthly_payment,
                'Principal': principal_payment,
                'Interest': interest_payment,
                'Ending_Balance': max(0, balance)
            })
            
            if balance <= 0:
                break
                
        return pd.DataFrame(schedule)
    
    def scenario_analysis(self, scenarios):
        """Analyze multiple refinancing scenarios"""
        results = []
        
        # Current scenario
        current_apt_payment = self.calculate_monthly_payment(self.apartment_loan, self.apartment_rate, self.apartment_term)
        current_inv_payment = self.calculate_monthly_payment(self.investment_loan, self.investment_rate, self.investment_term)
        current_total = current_apt_payment + current_inv_payment
        
        results.append({
            'Scenario': 'Current',
            'Apt_Rate': self.apartment_rate,
            'Inv_Rate': self.investment_rate,
            'Apt_Payment': current_apt_payment,
            'Inv_Payment': current_inv_payment,
            'Total_Payment': current_total,
            'Monthly_Savings': 0,
            'Annual_Savings': 0,
            'Refi_Costs': 0,
            'Breakeven_Months': 'N/A'
        })
        
        # Scenario analysis
        for i, scenario in enumerate(scenarios, 1):
            apt_rate = scenario.get('apt_rate', self.apartment_rate)
            inv_rate = scenario.get('inv_rate', self.investment_rate)
            apt_term = scenario.get('apt_term', self.apartment_term)
            inv_term = scenario.get('inv_term', self.investment_term)
            refi_costs = scenario.get('refi_costs', 5000)
            
            new_apt_payment = self.calculate_monthly_payment(self.apartment_loan, apt_rate, apt_term)
            new_inv_payment = self.calculate_monthly_payment(self.investment_loan, inv_rate, inv_term)
            new_total = new_apt_payment + new_inv_payment
            
            monthly_savings = current_total - new_total
            annual_savings = monthly_savings * 12
            breakeven = math.ceil(refi_costs / monthly_savings) if monthly_savings > 0 else 'N/A'
            
            results.append({
                'Scenario': f'Scenario {i}',
                'Apt_Rate': apt_rate,
                'Inv_Rate': inv_rate,
                'Apt_Payment': new_apt_payment,
                'Inv_Payment': new_inv_payment,
                'Total_Payment': new_total,
                'Monthly_Savings': monthly_savings,
                'Annual_Savings': annual_savings,
                'Refi_Costs': refi_costs,
                'Breakeven_Months': breakeven
            })
        
        return pd.DataFrame(results)
    
    def create_excel_model(self, filename='Beautiful_Loan_Scenario_Analysis.xlsx'):
        """Create comprehensive Excel model with beautiful styling"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            print("Creating styles...")
            # Define styles
            self._create_styles(wb)
            
            print("Creating dashboard...")
            # Create all worksheets
            self._create_dashboard(wb)
            
            print("Creating apartment analysis...")
            self._create_apartment_analysis(wb)
            
            print("Creating investment analysis...")
            self._create_investment_analysis(wb)
            
            print("Creating scenario comparison...")
            self._create_scenario_comparison(wb)
            
            print("Creating amortization schedules...")
            self._create_amortization_schedules(wb)
            
            print("Saving workbook...")
            # Save workbook
            wb.save(filename)
            print(f"Excel model created: {filename}")
            return filename
            
        except Exception as e:
            print(f"Detailed error in create_excel_model: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _create_styles(self, wb):
        """Create custom styles for the workbook"""
        # Title style
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(name='Segoe UI', size=18, bold=True, color='2C3E50')
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        wb.add_named_style(title_style)
        
        # Header style (blue gradient)
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='3498DB', end_color='2E86AB', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        wb.add_named_style(header_style)
        
        # Subheader style (darker blue)
        subheader_style = NamedStyle(name="subheader_style")
        subheader_style.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        subheader_style.fill = PatternFill(start_color='2C3E50', end_color='34495E', fill_type='solid')
        subheader_style.alignment = Alignment(horizontal='left', vertical='center')
        wb.add_named_style(subheader_style)
        
        # Data style
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(name='Segoe UI', size=10)
        data_style.alignment = Alignment(horizontal='right', vertical='center')
        wb.add_named_style(data_style)
        
        # Currency style
        currency_style = NamedStyle(name="currency_style")
        currency_style.font = Font(name='Segoe UI', size=10)
        currency_style.alignment = Alignment(horizontal='right', vertical='center')
        currency_style.number_format = '$#,##0'
        wb.add_named_style(currency_style)
        
        # Percentage style
        percentage_style = NamedStyle(name="percentage_style")
        percentage_style.font = Font(name='Segoe UI', size=10)
        percentage_style.alignment = Alignment(horizontal='right', vertical='center')
        percentage_style.number_format = '0.00%'
        wb.add_named_style(percentage_style)
        
        # Positive style (green for savings)
        positive_style = NamedStyle(name="positive_style")
        positive_style.font = Font(name='Segoe UI', size=10, bold=True, color='27AE60')
        positive_style.alignment = Alignment(horizontal='right', vertical='center')
        positive_style.fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
        positive_style.number_format = '$#,##0'
        wb.add_named_style(positive_style)
        
        # Summary card style
        summary_style = NamedStyle(name="summary_style")
        summary_style.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
        summary_style.fill = PatternFill(start_color='8E44AD', end_color='9B59B6', fill_type='solid')
        summary_style.alignment = Alignment(horizontal='center', vertical='center')
        wb.add_named_style(summary_style)
    
    def _create_dashboard(self, wb):
        """Create beautifully styled dashboard worksheet"""
        ws = wb.create_sheet("🏠 Dashboard")
        
        # Set row heights and column widths
        ws.row_dimensions[1].height = 30
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 3  # Spacer
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 18
        
        # Main title
        ws['A1'] = "💼 Loan Scenario Analysis Dashboard"
        ws['A1'].style = "title_style"
        ws.merge_cells('A1:E1')
        
        # Summary cards section
        ws['A3'] = "📊 PORTFOLIO OVERVIEW"
        ws['A3'].style = "header_style"
        ws.merge_cells('A3:E3')
        
        # Summary cards
        ws['A5'] = "Total Loan Amount"
        ws['A5'].style = "subheader_style"
        ws['B5'] = f"=B11+B16"
        ws['B5'].style = "summary_style"
        ws['B5'].number_format = '$#,##0'
        
        ws['D5'] = "Total Monthly Payment"
        ws['D5'].style = "subheader_style"
        ws['E5'] = f"=B14+B19"
        ws['E5'].style = "summary_style"
        ws['E5'].number_format = '$#,##0'
        
        # Current loans section
        ws['A7'] = "🏡 CURRENT LOAN DETAILS"
        ws['A7'].style = "header_style"
        ws.merge_cells('A7:E7')
        
        # Apartment loan
        ws['A9'] = "🏢 Apartment Loan"
        ws['A9'].style = "subheader_style"
        ws.merge_cells('A9:B9')
        
        ws['A11'] = "Principal Amount:"
        ws['B11'] = self.apartment_loan
        ws['A12'] = "Interest Rate:"
        ws['B12'] = self.apartment_rate/100
        ws['A13'] = "Term Remaining (years):"
        ws['B13'] = self.apartment_term
        ws['A14'] = "Monthly Payment:"
        ws['B14'] = f"=PMT(B12/12,B13*12,B11)"
        
        # Investment loan
        ws['A16'] = "🏘️ Investment Property"
        ws['A16'].style = "subheader_style"
        ws.merge_cells('A16:B16')
        
        ws['A17'] = "Principal Amount:"
        ws['B17'] = self.investment_loan
        ws['A18'] = "Interest Rate:"
        ws['B18'] = self.investment_rate/100
        ws['A19'] = "Term Remaining (years):"
        ws['B19'] = self.investment_term
        ws['A20'] = "Monthly Payment:"
        ws['B20'] = f"=PMT(B18/12,B19*12,B17)"
        
        # Style the data cells
        for row in range(11, 21):
            ws[f'A{row}'].style = "data_style"
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
            cell_value = ws[f'A{row}'].value
            if cell_value and ("Amount" in str(cell_value) or "Payment" in str(cell_value)):
                ws[f'B{row}'].style = "currency_style"
            elif cell_value and "Rate" in str(cell_value):
                ws[f'B{row}'].style = "percentage_style"
            else:
                ws[f'B{row}'].style = "data_style"
        
        # What-if scenario section
        ws['A22'] = "🎯 WHAT-IF SCENARIO ANALYZER"
        ws['A22'].style = "header_style"
        ws.merge_cells('A22:E22')
        
        ws['A24'] = "New Apartment Rate:"
        ws['B24'] = 4.99/100
        ws['A25'] = "New Investment Rate:"
        ws['B25'] = 5.99/100
        ws['A26'] = "Refinance Costs:"
        ws['B26'] = 5000
        
        ws['A28'] = "New Total Payment:"
        ws['B28'] = f"=PMT(B24/12,B13*12,B11)+PMT(B25/12,B19*12,B17)"
        ws['A29'] = "Monthly Savings:"
        ws['B29'] = f"=B14+B20-B28"
        ws['A30'] = "Break-even (months):"
        ws['B30'] = f"=IF(B29>0,B26/B29,\"N/A\")"
        
        # Style the scenario cells
        for row in range(24, 31):
            ws[f'A{row}'].style = "data_style"
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
            cell_value = ws[f'A{row}'].value
            if cell_value and "Rate" in str(cell_value):
                ws[f'B{row}'].style = "percentage_style"
                ws[f'B{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            elif cell_value and "Costs" in str(cell_value):
                ws[f'B{row}'].style = "currency_style"
                ws[f'B{row}'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            elif cell_value and "Savings" in str(cell_value):
                ws[f'B{row}'].style = "positive_style"
            else:
                ws[f'B{row}'].style = "currency_style"
    
    def _create_apartment_analysis(self, wb):
        """Create apartment loan analysis worksheet"""
        ws = wb.create_sheet("🏢 Apartment Loan")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        
        # Title
        ws['A1'] = "🏢 APARTMENT LOAN ANALYSIS"
        ws['A1'].style = "title_style"
        ws.merge_cells('A1:E1')
        
        # Current loan details
        ws['A3'] = "📋 Current Loan Details"
        ws['A3'].style = "header_style"
        ws.merge_cells('A3:E3')
        
        current_data = [
            ("Principal Amount:", self.apartment_loan, "A5", "B5"),
            ("Interest Rate:", self.apartment_rate/100, "A6", "B6"),
            ("Term Remaining (years):", self.apartment_term, "A7", "B7"),
            ("Monthly Payment:", "=PMT(B6/12,B7*12,B5)", "A8", "B8"),
            ("Total Interest:", "=B8*B7*12-B5", "A9", "B9"),
        ]
        
        for label, value, label_cell, value_cell in current_data:
            ws[label_cell] = label
            ws[label_cell].style = "data_style"
            ws[label_cell].alignment = Alignment(horizontal='left')
            ws[value_cell] = value
            if "Rate" in label:
                ws[value_cell].style = "percentage_style"
            else:
                ws[value_cell].style = "currency_style"
        
        # Refinance scenario
        ws['A11'] = "🔄 Refinance Scenario"
        ws['A11'].style = "header_style"
        ws.merge_cells('A11:E11')
        
        refi_data = [
            ("New Interest Rate:", 4.99/100, "A13", "B13"),
            ("New Term (years):", 5, "A14", "B14"),
            ("Refinance Costs:", 3000, "A15", "B15"),
            ("New Monthly Payment:", "=PMT(B13/12,B14*12,B5)", "A16", "B16"),
            ("New Total Interest:", "=B16*B14*12-B5", "A17", "B17"),
        ]
        
        for label, value, label_cell, value_cell in refi_data:
            ws[label_cell] = label
            ws[label_cell].style = "data_style"
            ws[label_cell].alignment = Alignment(horizontal='left')
            ws[value_cell] = value
            if "Rate" in label:
                ws[value_cell].style = "percentage_style"
            else:
                ws[value_cell].style = "currency_style"
        
        # Comparison table
        ws['A19'] = "⚖️ COMPARISON"
        ws['A19'].style = "header_style"
        ws.merge_cells('A19:E19')
        
        # Headers
        headers = ["Metric", "Current", "Refinanced", "Difference"]
        for i, header in enumerate(headers, 1):
            ws.cell(row=21, column=i, value=header).style = "subheader_style"
        
        # Comparison data
        comparison = [
            ("Monthly Payment", "=B8", "=B16", "=C22-B22"),
            ("Total Interest", "=B9", "=B17", "=C23-B23"),
            ("Net Savings", "", "=C23-B23-B15", "=C24"),
            ("Break-even (months)", "", "=IF(ABS(D22)>0,B15/ABS(D22),\"N/A\")", "=C25")
        ]
        
        for i, (metric, current, refinanced, difference) in enumerate(comparison, 22):
            ws[f'A{i}'] = metric
            ws[f'A{i}'].style = "data_style"
            ws[f'A{i}'].alignment = Alignment(horizontal='left')
            
            if current:
                ws[f'B{i}'] = current
                ws[f'B{i}'].style = "currency_style"
            if refinanced:
                ws[f'C{i}'] = refinanced
                ws[f'C{i}'].style = "currency_style"
            if difference:
                ws[f'D{i}'] = difference
                ws[f'D{i}'].style = "currency_style"
    
    def _create_investment_analysis(self, wb):
        """Create investment property analysis worksheet"""
        ws = wb.create_sheet("🏘️ Investment Property")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        
        # Title
        ws['A1'] = "🏘️ INVESTMENT PROPERTY ANALYSIS"
        ws['A1'].style = "title_style"
        ws.merge_cells('A1:E1')
        
        # Current loan details
        ws['A3'] = "📋 Current Loan Details"
        ws['A3'].style = "header_style"
        ws.merge_cells('A3:E3')
        
        current_data = [
            ("Principal Amount:", self.investment_loan, "A5", "B5"),
            ("Interest Rate:", self.investment_rate/100, "A6", "B6"),
            ("Term Remaining (years):", self.investment_term, "A7", "B7"),
            ("Monthly Payment:", "=PMT(B6/12,B7*12,B5)", "A8", "B8"),
            ("Total Interest:", "=B8*B7*12-B5", "A9", "B9"),
        ]
        
        for label, value, label_cell, value_cell in current_data:
            ws[label_cell] = label
            ws[label_cell].style = "data_style"
            ws[label_cell].alignment = Alignment(horizontal='left')
            ws[value_cell] = value
            if "Rate" in label:
                ws[value_cell].style = "percentage_style"
            else:
                ws[value_cell].style = "currency_style"
        
        # Refinance scenario
        ws['A11'] = "🔄 Refinance Scenario"
        ws['A11'].style = "header_style"
        ws.merge_cells('A11:E11')
        
        refi_data = [
            ("New Interest Rate:", 5.99/100, "A13", "B13"),
            ("New Term (years):", 30, "A14", "B14"),
            ("Refinance Costs:", 2000, "A15", "B15"),
            ("New Monthly Payment:", "=PMT(B13/12,B14*12,B5)", "A16", "B16"),
            ("New Total Interest:", "=B16*B14*12-B5", "A17", "B17"),
        ]
        
        for label, value, label_cell, value_cell in refi_data:
            ws[label_cell] = label
            ws[label_cell].style = "data_style"
            ws[label_cell].alignment = Alignment(horizontal='left')
            ws[value_cell] = value
            if "Rate" in label:
                ws[value_cell].style = "percentage_style"
            else:
                ws[value_cell].style = "currency_style"
        
        # Comparison table
        ws['A19'] = "⚖️ COMPARISON"
        ws['A19'].style = "header_style"
        ws.merge_cells('A19:E19')
        
        # Headers
        headers = ["Metric", "Current", "Refinanced", "Difference"]
        for i, header in enumerate(headers, 1):
            ws.cell(row=21, column=i, value=header).style = "subheader_style"
        
        # Comparison data
        comparison = [
            ("Monthly Payment", "=B8", "=B16", "=C22-B22"),
            ("Total Interest", "=B9", "=B17", "=C23-B23"),
            ("Net Savings", "", "=C23-B23-B15", "=C24"),
            ("Break-even (months)", "", "=IF(ABS(D22)>0,B15/ABS(D22),\"N/A\")", "=C25")
        ]
        
        for i, (metric, current, refinanced, difference) in enumerate(comparison, 22):
            ws[f'A{i}'] = metric
            ws[f'A{i}'].style = "data_style"
            ws[f'A{i}'].alignment = Alignment(horizontal='left')
            
            if current:
                ws[f'B{i}'] = current
                ws[f'B{i}'].style = "currency_style"
            if refinanced:
                ws[f'C{i}'] = refinanced
                ws[f'C{i}'].style = "currency_style"
            if difference:
                ws[f'D{i}'] = difference
                ws[f'D{i}'].style = "currency_style"
    
    def _create_scenario_comparison(self, wb):
        """Create scenario comparison worksheet"""
        ws = wb.create_sheet("📊 Scenario Comparison")
        
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 16
        
        # Title
        ws['A1'] = "📊 MULTI-SCENARIO COMPARISON"
        ws['A1'].style = "title_style"
        ws.merge_cells('A1:E1')
        
        # Input variables
        ws['A3'] = "🎛️ INPUT VARIABLES"
        ws['A3'].style = "header_style"
        ws.merge_cells('A3:E3')
        
        scenarios = ['Current', 'Scenario 1', 'Scenario 2', 'Scenario 3']
        for i, scenario in enumerate(scenarios):
            ws.cell(row=4, column=i+2, value=scenario).style = "subheader_style"
        
        # Input data
        input_data = [
            ("Apartment Rate (%)", [self.apartment_rate/100, 4.99/100, 5.25/100, 4.75/100]),
            ("Investment Rate (%)", [self.investment_rate/100, 5.99/100, 6.25/100, 5.75/100]),
            ("Refinance Costs", [0, 5000, 4000, 6000])
        ]
        
        for row_idx, (label, values) in enumerate(input_data, 5):
            ws[f'A{row_idx}'] = label
            ws[f'A{row_idx}'].style = "data_style"
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='left')
            
            for col_idx, value in enumerate(values, 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if "Rate" in label:
                    cell.style = "percentage_style"
                else:
                    cell.style = "currency_style"
        
        # Results section
        ws['A9'] = "📈 RESULTS"
        ws['A9'].style = "header_style"
        ws.merge_cells('A9:E9')
        
        for i, scenario in enumerate(scenarios):
            ws.cell(row=10, column=i+2, value=scenario).style = "subheader_style"
        
        # Calculations
        calc_data = [
            ("Apartment Payment", [f"=PMT(B5/12,{self.apartment_term}*12,{self.apartment_loan})",
                                  f"=PMT(C5/12,{self.apartment_term}*12,{self.apartment_loan})",
                                  f"=PMT(D5/12,{self.apartment_term}*12,{self.apartment_loan})",
                                  f"=PMT(E5/12,{self.apartment_term}*12,{self.apartment_loan})"]),
            ("Investment Payment", [f"=PMT(B6/12,{self.investment_term}*12,{self.investment_loan})",
                                   f"=PMT(C6/12,{self.investment_term}*12,{self.investment_loan})",
                                   f"=PMT(D6/12,{self.investment_term}*12,{self.investment_loan})",
                                   f"=PMT(E6/12,{self.investment_term}*12,{self.investment_loan})"]),
            ("Total Payment", ["=B11+B12", "=C11+C12", "=D11+D12", "=E11+E12"]),
            ("Monthly Savings", ["", "=B13-C13", "=B13-D13", "=B13-E13"]),
            ("Annual Savings", ["", "=C14*12", "=D14*12", "=E14*12"])
        ]
        
        for row_idx, (label, formulas) in enumerate(calc_data, 11):
            ws[f'A{row_idx}'] = label
            ws[f'A{row_idx}'].style = "data_style"
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='left')
            
            for col_idx, formula in enumerate(formulas, 2):
                if formula:
                    cell = ws.cell(row=row_idx, column=col_idx, value=formula)
                    if "Savings" in label and col_idx > 2:
                        cell.style = "positive_style"
                    else:
                        cell.style = "currency_style"
    
    def _create_amortization_schedules(self, wb):
        """Create amortization schedules"""
        # Apartment amortization
        ws_apt = wb.create_sheet("📋 Apt Amortization")
        
        ws_apt['A1'] = "📋 APARTMENT AMORTIZATION SCHEDULE"
        ws_apt['A1'].style = "title_style"
        ws_apt.merge_cells('A1:G1')
        
        ws_apt['A3'] = f"Loan: ${self.apartment_loan:,.0f} at {self.apartment_rate}% for {self.apartment_term} years"
        ws_apt['A3'].font = Font(name='Segoe UI', size=12, bold=True)
        
        # Create schedule
        apt_schedule = self.create_amortization_schedule(
            self.apartment_loan, self.apartment_rate, self.apartment_term
        )
        
        # Headers
        headers = ['Payment', 'Beginning Balance', 'Payment Amount', 'Principal', 'Interest', 'Ending Balance']
        for i, header in enumerate(headers, 1):
            ws_apt.cell(row=5, column=i, value=header).style = "subheader_style"
        
        # Add data
        for idx, row in apt_schedule.iterrows():
            row_num = idx + 6
            ws_apt[f'A{row_num}'] = row['Payment']
            ws_apt[f'B{row_num}'] = row['Beginning_Balance']
            ws_apt[f'C{row_num}'] = row['Payment_Amount']
            ws_apt[f'D{row_num}'] = row['Principal']
            ws_apt[f'E{row_num}'] = row['Interest']
            ws_apt[f'F{row_num}'] = row['Ending_Balance']
            
            # Style cells
            ws_apt[f'A{row_num}'].style = "data_style"
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_apt[f'{col}{row_num}'].style = "currency_style"
        
        # Set column widths
        column_widths = [10, 18, 15, 15, 15, 18]
        for i, width in enumerate(column_widths, 1):
            ws_apt.column_dimensions[chr(64 + i)].width = width
        
        # Investment amortization (abbreviated)
        ws_inv = wb.create_sheet("📋 Inv Amortization")
        
        ws_inv['A1'] = "📋 INVESTMENT PROPERTY AMORTIZATION SCHEDULE"
        ws_inv['A1'].style = "title_style"
        ws_inv.merge_cells('A1:G1')
        
        ws_inv['A3'] = f"Loan: ${self.investment_loan:,.0f} at {self.investment_rate}% for {self.investment_term} years"
        ws_inv['A3'].font = Font(name='Segoe UI', size=12, bold=True)
        
        # Create schedule
        inv_schedule = self.create_amortization_schedule(
            self.investment_loan, self.investment_rate, self.investment_term
        )
        
        # Headers
        for i, header in enumerate(headers, 1):
            ws_inv.cell(row=5, column=i, value=header).style = "subheader_style"
        
        # Add first 24 payments and then every 12th payment
        display_rows = []
        for idx, row in inv_schedule.iterrows():
            if idx < 24 or (idx + 1) % 12 == 0:
                display_rows.append((idx, row))
        
        for display_idx, (original_idx, row) in enumerate(display_rows):
            row_num = display_idx + 6
            ws_inv[f'A{row_num}'] = row['Payment']
            ws_inv[f'B{row_num}'] = row['Beginning_Balance']
            ws_inv[f'C{row_num}'] = row['Payment_Amount']
            ws_inv[f'D{row_num}'] = row['Principal']
            ws_inv[f'E{row_num}'] = row['Interest']
            ws_inv[f'F{row_num}'] = row['Ending_Balance']
            
            # Style cells
            ws_inv[f'A{row_num}'].style = "data_style"
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_inv[f'{col}{row_num}'].style = "currency_style"
        
        # Set column widths
        for i, width in enumerate(column_widths, 1):
            ws_inv.column_dimensions[chr(64 + i)].width = width
        
        # Add note
        note_row = len(display_rows) + 8
        ws_inv[f'A{note_row}'] = "Note: Schedule shows first 24 payments, then every 12th payment"
        ws_inv[f'A{note_row}'].font = Font(name='Segoe UI', size=9, italic=True)


def main():
    """Main function to create the loan analysis model"""
    print("🏗️  Creating beautiful Excel loan analysis model...")
    
    # Initialize analyzer with your loan details
    analyzer = LoanScenarioAnalyzer(
        apartment_loan=555000,
        apartment_rate=5.49,
        apartment_term=3,
        investment_loan=220000,
        investment_rate=6.49,
        investment_term=27.5
    )
    
    # Create Excel model
    filename = analyzer.create_excel_model('Beautiful_Loan_Scenario_Analysis.xlsx')
    
    # Run scenario analysis
    scenarios = [
        {'apt_rate': 4.99, 'inv_rate': 5.99, 'refi_costs': 5000},
        {'apt_rate': 5.25, 'inv_rate': 6.25, 'refi_costs': 4000},
        {'apt_rate': 4.75, 'inv_rate': 5.75, 'refi_costs': 6000}
    ]
    
    scenario_results = analyzer.scenario_analysis(scenarios)
    
    print("\n📊 Scenario Analysis Results:")
    print("=" * 80)
    for _, row in scenario_results.iterrows():
        print(f"{row['Scenario']:12} | Monthly: ${row['Monthly_Savings']:7.0f} | Annual: ${row['Annual_Savings']:8.0f} | Break-even: {row['Breakeven_Months']}")
    
    print(f"\n✅ Excel model created successfully: {filename}")
    print("\n🎨 Features included:")
    print("   • 🏠 Interactive Dashboard with summary cards")
    print("   • 🏢 Detailed Apartment Loan Analysis")
    print("   • 🏘️ Detailed Investment Property Analysis") 
    print("   • 📊 Multi-Scenario Comparison")
    print("   • 📋 Complete Amortization Schedules")
    print("   • 💎 Professional styling with colors and emojis")
    print("   • 🎯 Interactive what-if scenario calculator")
    
    print(f"\n🚀 Ready to analyze your refinancing options!")
    return filename


if __name__ == "__main__":
    # Required packages: pip install pandas openpyxl numpy
    
    try:
        filename = main()
        print(f"\n🎉 Success! Open '{filename}' to start your loan analysis!")
        print("\n💡 Tips for using your Excel model:")
        print("   • Modify the blue input cells in the Dashboard to test scenarios")
        print("   • Check the comparison tables for break-even analysis")
        print("   • Use the Scenario Comparison tab for side-by-side analysis")
        print("   • Review amortization schedules for detailed payment breakdowns")
        
    except ImportError as e:
        print("❌ Missing required package. Please install dependencies:")
        print("   pip install pandas openpyxl numpy")
        print(f"   Error: {e}")
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        print("Please check your file permissions and try again.")
        print("Make sure you have write access to the current directory.")